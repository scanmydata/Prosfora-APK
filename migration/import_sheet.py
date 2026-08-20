#!/usr/bin/env python3
"""
Μετατρέπει το export του Google Sheet (ΔΕΔΟΜΕΝΑ.xlsx) σε seed.json που
διαβάζει το app κατά την εισαγωγή δεδομένων.

    python migration/import_sheet.py [--dry-run]

Τα IDs του AppSheet διατηρούνται αυτούσια ως primary keys, οπότε οι σχέσεις
parent/child μεταφέρονται χωρίς αντιστοίχιση.
"""
from __future__ import annotations

import argparse
import json
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
SOURCE = ROOT / "source" / "ΔΕΔΟΜΕΝΑ.xlsx"
OUTPUT = ROOT.parent / "app" / "src" / "main" / "assets" / "seed.json"

STATUS_MAP = {
    "Δημιουργήθηκε": "CREATED",
    "Σε επεξεργασία": "IN_PROGRESS",
    "Ολοκληρώθηκε": "COMPLETED",
}


def epoch_day(value) -> int:
    if isinstance(value, datetime):
        value = value.date()
    if isinstance(value, date):
        return (value - date(1970, 1, 1)).days
    if isinstance(value, str) and value.strip():
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y"):
            try:
                return (datetime.strptime(value.strip(), fmt).date() - date(1970, 1, 1)).days
            except ValueError:
                continue
    return (date.today() - date(1970, 1, 1)).days


def number(value) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("€", "").strip()
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def rows(sheet):
    """Γραμμές με περιεχόμενο — το AppSheet αφήνει κενές γραμμές ανάμεσα."""
    header = [str(c.value).strip() if c.value else "" for c in next(sheet.iter_rows(max_row=1))]
    for raw in sheet.iter_rows(min_row=2, values_only=True):
        if not any(v not in (None, "") for v in raw):
            continue
        record = dict(zip(header, raw))
        if not str(record.get(header[0]) or "").strip():
            continue
        yield record


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true", help="τύπωσε τι θα γραφτεί, χωρίς εγγραφή")
    args = parser.parse_args()

    if not SOURCE.exists():
        print(f"Δεν βρέθηκε το {SOURCE}", file=sys.stderr)
        return 1

    book = openpyxl.load_workbook(SOURCE, data_only=True)

    offers = []
    for record in rows(book["Προσφορές"]):
        offers.append({
            "id": str(record["ID_Προσφοράς"]).strip(),
            "address": str(record.get("Οδός / Περιοχή") or "").strip(),
            "dateEpochDay": epoch_day(record.get("Ημερομηνία")),
            "kind": str(record.get("Είδος") or "").strip(),
            "email": str(record.get("Email") or "").strip(),
            "status": STATUS_MAP.get(str(record.get("Κατάσταση") or "").strip(), "CREATED"),
        })

    known_offers = {o["id"] for o in offers}

    spaces, orphan_spaces = [], []
    for i, record in enumerate(rows(book["Χώροι_έργου"])):
        offer_id = str(record.get("ID_Προσφοράς") or "").strip()
        entry = {
            "id": str(record["ID_Χώρου"]).strip(),
            "offerId": offer_id,
            "description": str(record.get("Περιγραφή Χώρου") or "").strip(),
            "area": number(record.get("Επιφάνεια (τ.μ.)")),
            "unitPrice": number(record.get("Τιμή Μονάδος")),
            "position": i,
        }
        (spaces if offer_id in known_offers else orphan_spaces).append(entry)

    notes, orphan_notes = [], []
    for i, record in enumerate(rows(book["Λίστα_Παρατηρήσεων"])):
        offer_id = str(record.get("ID_Προσφοράς") or "").strip()
        entry = {
            "id": str(record["ID_Παρατήρησης"]).strip(),
            "offerId": offer_id,
            "text": str(record.get("Κείμενο") or "").strip(),
            "position": i,
        }
        (notes if offer_id in known_offers else orphan_notes).append(entry)

    # Οι σημειώσεις που επαναλαμβάνονται γίνονται έτοιμες επιλογές στο app
    counts: dict[str, int] = {}
    for note in notes:
        counts[note["text"]] = counts.get(note["text"], 0) + 1
    presets = [
        {"text": text, "position": i, "useCount": count}
        for i, (text, count) in enumerate(
            sorted(counts.items(), key=lambda kv: (-kv[1], kv[0])),
        )
        if count > 1 or len(text) < 200
    ]

    payload = {"offers": offers, "spaces": spaces, "notes": notes, "presets": presets}

    print(f"Προσφορές: {len(offers)}")
    print(f"Χώροι:     {len(spaces)}" + (f"  ⚠️ {len(orphan_spaces)} ορφανοί" if orphan_spaces else ""))
    print(f"Σημειώσεις:{len(notes)}" + (f"  ⚠️ {len(orphan_notes)} ορφανές" if orphan_notes else ""))
    print(f"Έτοιμες σημειώσεις: {len(presets)}")
    for offer in offers:
        # ίδια σειρά πράξεων με το app: στρογγυλοποίηση ανά γραμμή, μετά άθροισμα
        total = sum(round(s["area"] * s["unitPrice"], 2) for s in spaces if s["offerId"] == offer["id"])
        n = sum(1 for s in spaces if s["offerId"] == offer["id"])
        print(f"  · {offer['address']}: {n} χώροι, σύνολο {total:.2f} €")

    if orphan_spaces or orphan_notes:
        print("\n⚠️ Οι ορφανές εγγραφές δεν μεταφέρονται (δείχνουν σε προσφορά που δεν υπάρχει):")
        for entry in orphan_spaces + orphan_notes:
            print(f"   {entry['id']} → ID_Προσφοράς={entry['offerId'] or '(κενό)'}")

    if args.dry_run:
        print("\n(dry run — δεν γράφτηκε τίποτα)")
        return 0

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\nΓράφτηκε: {OUTPUT.relative_to(ROOT.parent)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
